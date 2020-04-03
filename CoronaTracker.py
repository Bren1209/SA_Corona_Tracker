from bs4 import BeautifulSoup as bs
import requests
import tkinter as tk
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import os.path
from datetime import date

##################################################################
'''Error catching functions'''
##################################################################


def conn_err():
    err_root1 = tk.Tk()

    err_root1.iconbitmap('./covid_tkinter_icon.ico')
    err_root1.title('SA Covid-19')
    err_root1.geometry('250x225+550+200')

    canvas2 = tk.Canvas(err_root1, width=250, height=225)
    canvas2.pack()

    conn_err_label = tk.Label(text='Connection Error\n---\nPlease check your internet connection\nand try again.', font='Arial 9 bold')
    ok_button = tk.Button(text='OK', command=err_root1.quit)

    canvas2.create_window(125, 100, window=conn_err_label)
    canvas2.create_window(125, 200, window=ok_button)

    err_root1.mainloop()


def html_err():
    err_root2 = tk.Tk()

    err_root2.iconbitmap('./covid_tkinter_icon.ico')
    err_root2.title('SA Covid-19')
    err_root2.geometry('250x225+550+200')

    canvas2 = tk.Canvas(err_root2, width=250, height=225)
    canvas2.pack()

    conn_err_label = tk.Label(text='Index Error\n---\nPossibly due to a change\nin the website\'s structure.', font='Arial 9 bold')
    ok_button = tk.Button(text='OK', command=err_root2.quit)

    canvas2.create_window(125, 100, window=conn_err_label)
    canvas2.create_window(125, 200, window=ok_button)

    err_root2.mainloop()


##################################################################
'''Get the info from Worldometers, creates and modifies a list and 
    adds the data to relevant variables for later use.'''
##################################################################

site = 'https://www.worldometers.info/coronavirus/'

try:
    requests.get(site)
except requests.exceptions.ConnectionError:
    conn_err()

url = requests.get(site).content

soup = bs(url, 'lxml')

proper_string = ''

for item in soup.find_all('tr'):
    if 'South Africa' in item.text:
        proper_string += item.text

removed_newlines = proper_string.replace('\n', ' ')
create_string = removed_newlines.split(' ')

# print(create_string)

# while '' in create_string:
#     create_string.remove('')

try:
    int(create_string[3].replace(",", ""))
except IndexError:
    html_err()
except ValueError:
    html_err()

cases = int(create_string[3].replace(",", ""))
new_cases = create_string[4]  # Might error out due to plus sign, but not present in list?
deaths = int(create_string[5])
new_deaths = create_string[7]
recoveries = int(create_string[8])
today = date.today().strftime('%d/%m')

if new_cases == '':
    new_cases = 'Pending'
else:
    new_cases = int(create_string[4])

if new_deaths == '':
    new_deaths = 'Pending'
else:
    new_deaths = int(create_string[7])

#################################################################
'''Creates an xlsx sheet with data (history included) or adds to
    an existing sheet. Also plots the data on a graph'''
#################################################################


def write_and_plot():

    wb2 = load_workbook('Case_History.xlsx')
    worksheet2 = wb2['Case_History']
    new_row = worksheet2.max_row + 1
    last_row_deets = []
    new_deets = [cases, recoveries, deaths, today]

    for row in worksheet2.iter_rows(min_row=worksheet2.max_row, max_col=4, max_row=worksheet2.max_row):
        for cell in row:
            last_row_deets.append(cell.value)

    if last_row_deets != new_deets:

        worksheet2.cell(column=1, row=new_row, value=cases)
        worksheet2.cell(column=2, row=new_row, value=recoveries)
        worksheet2.cell(column=3, row=new_row, value=deaths)
        worksheet2.cell(column=4, row=new_row, value=today)

        wb2.save('Case_History.xlsx')

    else:
        pass

    cases_data = []
    recoveries_data = []
    deaths_data = []
    dates = []

    for row in worksheet2.values:
        cases_data.append(row[0])
        recoveries_data.append(row[1])
        deaths_data.append(row[2])
        dates.append(row[3])

    date_x = dates[1:]
    recoveries_x = recoveries_data[1:]
    deaths_x = deaths_data[1:]
    case_y = cases_data[1:]

    plt.plot(date_x, case_y, label='Confirmed')
    plt.plot(recoveries_x, label='Recoveries')
    plt.plot(deaths_x, label='Deaths')
    plt.xticks(rotation=45)
    plt.tight_layout(pad=3.7, w_pad=2, h_pad=3.8)

    plt.xlabel('DATE')
    plt.ylabel('CASE COUNT')
    plt.legend()
    plt.suptitle('SA Covid-19 Case History')


if os.path.isfile('Case_History.xlsx'):

    write_and_plot()

else:

    wb = Workbook()

    worksheet = wb.active
    worksheet1 = wb.create_sheet('Case_History', 0)

    worksheet1['A1'] = 'CONFIRMED CASES'
    worksheet1['B1'] = 'RECOVERIES'
    worksheet1['C1'] = 'DEATHS'
    worksheet1['D1'] = 'DATE'

    wb.save('Case_History.xlsx')

    write_and_plot()


#################################################################
    '''Creates a GUI with the data and an exit button'''
#################################################################

root = tk.Tk()

root.iconbitmap('./covid_tkinter_icon.ico')
root.title('SA Covid-19')
root.geometry('250x225+550+200')

canvas1 = tk.Canvas(root, width=250, height=225)
canvas1.pack()

label1 = tk.Label(text=f'Confirmed Cases: {cases}', font='Arial 10 bold')
label2 = tk.Label(text=f'Recoveries: {recoveries}', font='Arial 10 bold')
label3 = tk.Label(text=f'Deaths: {deaths}', font='Arial 10 bold')
label4 = tk.Label(text=f'New Cases: {new_cases}', fg='red', font='Arial 10 bold')
label5 = tk.Label(text=f'New Deaths: {new_deaths}', fg='red', font='Arial 10 bold')
quit_button = tk.Button(text='Exit', command=root.quit)
graph_button = tk.Button(text='Show Graph', command=plt.show)
img = tk.PhotoImage(file='SA_flag_converted.PPM')
canvas1.create_image(-20, 70, anchor='nw', image=img)

label_list = [label1, label2, label3, label4, label5]
window_width = 125
window_height = 30

for label in label_list:
    canvas1.create_window(window_width, window_height, window=label)
    window_height += 25

canvas1.create_window(125, 200, window=quit_button)
canvas1.create_window(125, 165, window=graph_button)

root.mainloop()
